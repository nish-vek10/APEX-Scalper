# /scripts/04_generate_report.py
"""
APEX Scalper — Pipeline Step 4: Generate Excel Report
=======================================================
Loads backtest_results.pkl and generates a fully-formatted,
investor-presentable Excel workbook with 6 sheets:

    Sheet 1 — Summary Dashboard      Overall portfolio KPIs
    Sheet 2 — Instrument Performance  Side-by-side per-asset table
    Sheet 3 — Monthly Returns         Heatmap: Portfolio (rows=years, cols=months)
    Sheet 4 — Monthly Returns (Assets) Same heatmap per instrument
    Sheet 5 — Trade Log               Full trade-by-trade detail
    Sheet 6 — Equity Data             Raw time-series equity values

MUST run Step 2 first:
    python -m scripts.02_run_backtest

Run from project root:
    python -m scripts.04_generate_report

Output: backtests/reports/APEX_Scalper_Report_{date}.xlsx
"""

import os
import sys
import pickle
import numpy as np
import pandas as pd
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config.settings import RESULTS_DIR, REPORTS_DIR, INITIAL_CAPITAL
from config.instruments import INSTRUMENT_CONFIG
from src.utils.logger import get_logger

logger = get_logger(__name__)

# ─────────────────────────────────────────────
# CONFIGURABLE STYLE PALETTE
# ─────────────────────────────────────────────
C = {
    # Header / banner
    "header_bg":    "0D1B2A",   # Dark navy
    "header_text":  "FFFFFF",
    # Section titles
    "section_bg":   "1A3A5C",
    "section_text": "FFFFFF",
    # Sub-headers (column labels)
    "col_bg":       "1E4976",
    "col_text":     "FFFFFF",
    # Metric labels (left column)
    "label_bg":     "EEF3F8",
    "label_text":   "0D1B2A",
    # Alternating row tints
    "row_light":    "F7FAFD",
    "row_dark":     "E8EFF7",
    # Positive values
    "pos_bg":       "E8F5E9",
    "pos_text":     "1B5E20",
    # Negative values
    "neg_bg":       "FFEBEE",
    "neg_text":     "B71C1C",
    # Neutral / accent
    "gold":         "F4D03F",
    "border":       "B0BEC5",
    "white":        "FFFFFF",
    # Heatmap green → red
    "heat_max_g":   "00695C",   # Deep green
    "heat_mid_g":   "A5D6A7",   # Light green
    "heat_zero":    "FFFFFF",   # White
    "heat_mid_r":   "EF9A9A",   # Light red
    "heat_max_r":   "B71C1C",   # Deep red
}

FONT_NAME = "Arial"

# Month labels for heatmap
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


# ═══════════════════════════════════════════════════════════════════════════
# LOW-LEVEL FORMATTING HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def _font(size=10, bold=False, color="000000", italic=False):
    return Font(name=FONT_NAME, size=size, bold=bold,
                color=color, italic=italic)

def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin", color="B0BEC5"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom(color="1A3A5C"):
    s = Side(style="medium", color=color)
    return Border(bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def _merge_write(ws, cell_range, value, font=None, fill=None,
                 alignment=None, border=None, num_fmt=None):
    """Merge a cell range, write value, and apply formatting to the top-left cell."""
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    if num_fmt:   cell.number_format = num_fmt

def _write(ws, row, col, value, font=None, fill=None,
           alignment=None, border=None, num_fmt=None):
    """Write a single cell with optional formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    if num_fmt:   cell.number_format = num_fmt
    return cell

def _pnl_color(value: float) -> tuple:
    """Return (bg_color, text_color) based on positive/negative value."""
    if value > 0:   return C["pos_bg"], C["pos_text"]
    elif value < 0: return C["neg_bg"], C["neg_text"]
    return C["white"], "000000"

def _heat_color(value: float, max_abs: float) -> str:
    """
    Interpolate between deep-green (max positive) → white (zero) →
    deep-red (max negative). Returns hex string (no '#').
    """
    if max_abs == 0:
        return C["heat_zero"]
    ratio = max(min(value / max_abs, 1.0), -1.0)

    def hex_to_rgb(h): return int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
    def lerp(a, b, t): return int(a + (b - a) * t)

    if ratio >= 0:
        rgb0 = hex_to_rgb(C["heat_zero"])
        rgb1 = hex_to_rgb(C["heat_max_g"])
        t = ratio
    else:
        rgb0 = hex_to_rgb(C["heat_zero"])
        rgb1 = hex_to_rgb(C["heat_max_r"])
        t = abs(ratio)

    r = lerp(rgb0[0], rgb1[0], t)
    g = lerp(rgb0[1], rgb1[1], t)
    b = lerp(rgb0[2], rgb1[2], t)
    return f"{r:02X}{g:02X}{b:02X}"


# ═══════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════

def _build_summary(wb: Workbook, metrics: dict, trades_df: pd.DataFrame,
                   start_date: str, end_date: str):
    """
    Sheet 1 — Summary Dashboard
    Two-column KPI grid with section separators.
    """
    ws = wb.active
    ws.title = "1. Summary Dashboard"
    ws.sheet_view.showGridLines = False

    # ── Banner ────────────────────────────────────────────────────────────
    _merge_write(ws, "A1:H1",
                 "APEX SCALPER — BACKTEST RESULTS",
                 font=_font(18, bold=True, color=C["header_text"]),
                 fill=_fill(C["header_bg"]),
                 alignment=_align("center"))
    ws.row_dimensions[1].height = 36

    _merge_write(ws, "A2:H2",
                 f"Period: {start_date}  →  {end_date}   |   "
                 f"Initial Capital: ${INITIAL_CAPITAL:,.2f}   |   "
                 f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}",
                 font=_font(10, color=C["header_text"]),
                 fill=_fill(C["section_bg"]),
                 alignment=_align("center"))
    ws.row_dimensions[2].height = 20

    def section_header(row, label):
        _merge_write(ws, f"A{row}:H{row}", label,
                     font=_font(11, bold=True, color=C["section_text"]),
                     fill=_fill(C["section_bg"]),
                     alignment=_align("left"))
        ws.row_dimensions[row].height = 20

    def kpi_row(row, label, value, fmt="text", positive_good=True):
        """Write label + value pair with colour coding."""
        bg = C["row_light"] if row % 2 == 0 else C["row_dark"]
        _write(ws, row, 1, label,
               font=_font(10, bold=True, color=C["label_text"]),
               fill=_fill(C["label_bg"]),
               alignment=_align("left"),
               border=_border())
        ws.merge_cells(f"B{row}:D{row}")

        if fmt == "currency":
            bg_v, txt_v = _pnl_color(value) if positive_good else (bg, "000000")
            cell = ws.cell(row=row, column=2, value=value)
            cell.font      = _font(10, bold=True, color=txt_v)
            cell.fill      = _fill(bg_v)
            cell.alignment = _align("right")
            cell.border    = _border()
            cell.number_format = '#,##0.00'
        elif fmt == "pct":
            bg_v, txt_v = _pnl_color(value) if positive_good else (bg, "000000")
            cell = ws.cell(row=row, column=2, value=value / 100)
            cell.font      = _font(10, bold=True, color=txt_v)
            cell.fill      = _fill(bg_v)
            cell.alignment = _align("right")
            cell.border    = _border()
            cell.number_format = "0.00%"
        elif fmt == "ratio":
            txt_v = C["pos_text"] if value > 1 else C["neg_text"] if value < 0 else "000000"
            cell = ws.cell(row=row, column=2, value=value)
            cell.font      = _font(10, bold=True, color=txt_v)
            cell.fill      = _fill(bg)
            cell.alignment = _align("right")
            cell.border    = _border()
            cell.number_format = "0.00"
        else:
            cell = ws.cell(row=row, column=2, value=value)
            cell.font      = _font(10, color="000000")
            cell.fill      = _fill(bg)
            cell.alignment = _align("right")
            cell.border    = _border()

    r = 4
    section_header(r, "  PORTFOLIO OVERVIEW"); r += 1
    kpi_row(r, "Backtest Period",    f"{start_date}  →  {end_date}"); r += 1
    kpi_row(r, "Initial Capital",    INITIAL_CAPITAL,              fmt="currency"); r += 1
    kpi_row(r, "Final Equity",       metrics.get("final_equity", 0), fmt="currency"); r += 1
    kpi_row(r, "Net P&L ($)",        metrics.get("net_pnl", 0),    fmt="currency"); r += 1
    kpi_row(r, "Total Return (%)",   metrics.get("total_return_pct", 0), fmt="pct"); r += 1
    kpi_row(r, "Peak Equity",        metrics.get("peak_equity", 0), fmt="currency"); r += 1

    r += 1
    section_header(r, "  TRADE STATISTICS"); r += 1
    kpi_row(r, "Total Trades",       metrics.get("total_trades", 0)); r += 1
    kpi_row(r, "Winning Trades",     metrics.get("win_count", 0)); r += 1
    kpi_row(r, "Losing Trades",      metrics.get("loss_count", 0)); r += 1
    kpi_row(r, "Win Rate (%)",       metrics.get("win_rate_pct", 0), fmt="pct"); r += 1
    kpi_row(r, "Profit Factor",      metrics.get("profit_factor", 0), fmt="ratio"); r += 1
    kpi_row(r, "Avg Win ($)",        metrics.get("avg_win", 0),   fmt="currency"); r += 1
    kpi_row(r, "Avg Loss ($)",       metrics.get("avg_loss", 0),  fmt="currency"); r += 1
    kpi_row(r, "Avg R:R",            metrics.get("avg_rr", 0),    fmt="ratio"); r += 1
    kpi_row(r, "Total Commission ($)", metrics.get("total_commission", 0), fmt="currency"); r += 1

    r += 1
    section_header(r, "  RISK-ADJUSTED PERFORMANCE"); r += 1
    kpi_row(r, "Max Drawdown (%)",   metrics.get("max_drawdown_pct", 0), fmt="pct",
            positive_good=False); r += 1
    kpi_row(r, "Sharpe Ratio",       metrics.get("sharpe_ratio", 0),  fmt="ratio"); r += 1
    kpi_row(r, "Sortino Ratio",      metrics.get("sortino_ratio", 0), fmt="ratio"); r += 1
    kpi_row(r, "Calmar Ratio",       metrics.get("calmar_ratio", 0),  fmt="ratio"); r += 1
    kpi_row(r, "Gross Profit ($)",   metrics.get("gross_profit", 0),  fmt="currency"); r += 1
    kpi_row(r, "Gross Loss ($)",     metrics.get("gross_loss", 0),    fmt="currency"); r += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    for col in ["E","F","G","H"]:
        ws.column_dimensions[col].width = 12


def _build_instrument_performance(wb: Workbook, metrics: dict):
    """
    Sheet 2 — Per-Instrument Performance
    Side-by-side comparison table for all 5 instruments + portfolio total.
    """
    ws = wb.create_sheet("2. Instrument Performance")
    ws.sheet_view.showGridLines = False

    _merge_write(ws, "A1:H1",
                 "PER-INSTRUMENT PERFORMANCE BREAKDOWN",
                 font=_font(14, bold=True, color=C["header_text"]),
                 fill=_fill(C["header_bg"]),
                 alignment=_align("center"))
    ws.row_dimensions[1].height = 32

    instruments = list(INSTRUMENT_CONFIG.keys())
    headers = (["Metric"] +
               [INSTRUMENT_CONFIG[s]["display_name"] for s in instruments] +
               ["Portfolio Combined"])
    col_letters = [get_column_letter(i+1) for i in range(len(headers))]

    # Column headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = _font(10, bold=True, color=C["col_text"])
        cell.fill      = _fill(C["col_bg"])
        cell.alignment = _align("center")
        cell.border    = _border()
    ws.row_dimensions[2].height = 20

    METRICS = [
        ("Total Trades",       "total_trades",       "int",      True),
        ("Win Count",          "win_count",           "int",      True),
        ("Loss Count",         "loss_count",          "int",      False),
        ("Win Rate (%)",       "win_rate_pct",        "pct",      True),
        ("Profit Factor",      "profit_factor",       "ratio",    True),
        ("Net P&L ($)",        "net_pnl",             "currency", True),
        ("Gross Profit ($)",   "gross_profit",        "currency", True),
        ("Gross Loss ($)",     "gross_loss",          "currency", False),
        ("Avg Win ($)",        "avg_win",             "currency", True),
        ("Avg Loss ($)",       "avg_loss",            "currency", False),
        ("Avg R:R",            "avg_rr",              "ratio",    True),
        ("Total Return (%)",   "total_return_pct",    "pct",      True),
        ("Final Equity ($)",   "final_equity",        "currency", True),
        ("Peak Equity ($)",    "peak_equity",         "currency", True),
        ("Max Drawdown (%)",   "max_drawdown_pct",    "pct",      False),
        ("Sharpe Ratio",       "sharpe_ratio",        "ratio",    True),
        ("Sortino Ratio",      "sortino_ratio",       "ratio",    True),
        ("Calmar Ratio",       "calmar_ratio",        "ratio",    True),
        ("Total Commission ($)","total_commission",   "currency", False),
    ]

    per_inst = metrics.get("per_instrument", {})

    for row_offset, (label, key, fmt, pos_good) in enumerate(METRICS):
        row = row_offset + 3
        bg  = C["row_light"] if row % 2 == 0 else C["row_dark"]

        # Metric label
        ws.cell(row=row, column=1, value=label).font = _font(9, bold=True)
        ws.cell(row=row, column=1).fill      = _fill(C["label_bg"])
        ws.cell(row=row, column=1).alignment = _align("left")
        ws.cell(row=row, column=1).border    = _border()

        # Per-instrument values
        for col_idx, symbol in enumerate(instruments, 2):
            m   = per_inst.get(symbol, {})
            val = m.get(key, 0) or 0
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = _align("center")
            cell.border    = _border()

            if fmt == "currency":
                cell.number_format = '#,##0.00'
                bg_v, txt_v = _pnl_color(val) if pos_good and val != 0 else (bg, "000000")
            elif fmt == "pct":
                cell.value         = val / 100
                cell.number_format = "0.00%"
                bg_v, txt_v = _pnl_color(val) if pos_good and val != 0 else (bg, "000000")
            elif fmt == "ratio":
                cell.number_format = "0.00"
                bg_v = C["pos_bg"] if (pos_good and val > 1) else C["neg_bg"] if (pos_good and val < 1) else bg
                txt_v = "000000"
            else:
                bg_v, txt_v = bg, "000000"

            cell.fill = _fill(bg_v)
            cell.font = _font(9, color=txt_v)

        # Portfolio column
        val  = metrics.get(key, 0) or 0
        col  = len(instruments) + 2
        cell = ws.cell(row=row, column=col, value=val)
        cell.font      = _font(9, bold=True, color="000000")
        cell.fill      = _fill(C["row_light"])
        cell.alignment = _align("center")
        cell.border    = _border()
        if fmt == "currency": cell.number_format = '#,##0.00'
        elif fmt == "pct":
            cell.value         = val / 100
            cell.number_format = "0.00%"
        elif fmt == "ratio": cell.number_format = "0.00"

    # Column widths
    ws.column_dimensions["A"].width = 28
    for i in range(1, len(instruments) + 3):
        ws.column_dimensions[get_column_letter(i+1)].width = 22


def _build_monthly_heatmap(wb: Workbook, trades_df: pd.DataFrame,
                            metrics: dict, sheet_num: int,
                            sheet_title: str, symbol: str = None):
    """
    Build a monthly returns heatmap sheet.
    If symbol is None → portfolio combined.
    Rows = calendar years. Columns = Jan … Dec + Annual Total.
    Cells are colour-coded green (profit) → white (zero) → red (loss).
    """
    ws = wb.create_sheet(sheet_title)
    ws.sheet_view.showGridLines = False

    title = (INSTRUMENT_CONFIG[symbol]["display_name"]
             if symbol else "Portfolio Combined")

    _merge_write(ws, "A1:O1",
                 f"MONTHLY RETURNS — {title.upper()}",
                 font=_font(13, bold=True, color=C["header_text"]),
                 fill=_fill(C["header_bg"]),
                 alignment=_align("center"))
    ws.row_dimensions[1].height = 30

    # Filter trades
    df = trades_df.copy()
    if symbol:
        df = df[df["instrument"] == symbol]

    df["open_time"] = pd.to_datetime(df["open_time"])
    df["year"]      = df["open_time"].dt.year
    df["month"]     = df["open_time"].dt.month

    # Pivot: rows=year, cols=month (1-12)
    monthly = (df.groupby(["year", "month"])["net_pnl"]
                 .sum()
                 .reset_index())
    pivot = monthly.pivot(index="year", columns="month", values="net_pnl").fillna(0)

    if pivot.empty:
        ws["A2"] = "No data available."
        return

    all_years = sorted(pivot.index.tolist())
    all_values = pivot.values.flatten()
    max_abs = max(abs(all_values.max()), abs(all_values.min())) if len(all_values) else 1

    # Column headers
    header_row = 3
    ws.cell(row=header_row, column=1, value="Year").font = _font(10, bold=True, color=C["col_text"])
    ws.cell(row=header_row, column=1).fill      = _fill(C["col_bg"])
    ws.cell(row=header_row, column=1).alignment = _align("center")
    ws.cell(row=header_row, column=1).border    = _border()

    for m_idx, m_name in enumerate(MONTHS, 2):
        cell = ws.cell(row=header_row, column=m_idx, value=m_name)
        cell.font      = _font(10, bold=True, color=C["col_text"])
        cell.fill      = _fill(C["col_bg"])
        cell.alignment = _align("center")
        cell.border    = _border()

    # Annual total header
    ws.cell(row=header_row, column=14, value="Annual P&L").font = _font(10, bold=True, color=C["section_text"])
    ws.cell(row=header_row, column=14).fill      = _fill(C["section_bg"])
    ws.cell(row=header_row, column=14).alignment = _align("center")
    ws.cell(row=header_row, column=14).border    = _border()

    ws.cell(row=header_row, column=15, value="Annual Ret %").font = _font(10, bold=True, color=C["section_text"])
    ws.cell(row=header_row, column=15).fill      = _fill(C["section_bg"])
    ws.cell(row=header_row, column=15).alignment = _align("center")
    ws.cell(row=header_row, column=15).border    = _border()

    ws.row_dimensions[header_row].height = 20

    # Data rows
    for yr_offset, year in enumerate(all_years):
        row = header_row + 1 + yr_offset
        ws.cell(row=row, column=1, value=year).font      = _font(10, bold=True)
        ws.cell(row=row, column=1).fill      = _fill(C["label_bg"])
        ws.cell(row=row, column=1).alignment = _align("center")
        ws.cell(row=row, column=1).border    = _border()

        annual_pnl = 0.0
        for m_num in range(1, 13):
            col  = m_num + 1
            val  = pivot.loc[year, m_num] if m_num in pivot.columns else 0.0
            val  = float(val) if not pd.isna(val) else 0.0
            annual_pnl += val

            hex_bg = _heat_color(val, max_abs)
            txt_color = "FFFFFF" if abs(val / max_abs) > 0.65 else "000000" if max_abs > 0 else "000000"

            cell = ws.cell(row=row, column=col, value=val)
            cell.font          = _font(9, bold=(val != 0), color=txt_color)
            cell.fill          = _fill(hex_bg)
            cell.alignment     = _align("center")
            cell.border        = _border()
            cell.number_format = '#,##0.00'

        # Annual P&L
        ann_bg, ann_txt = _pnl_color(annual_pnl)
        cell14 = ws.cell(row=row, column=14, value=annual_pnl)
        cell14.font          = _font(9, bold=True, color=ann_txt)
        cell14.fill          = _fill(ann_bg)
        cell14.alignment     = _align("center")
        cell14.border        = _border()
        cell14.number_format = '#,##0.00'

        # Annual Return %
        ann_ret = annual_pnl / INITIAL_CAPITAL * 100
        cell15  = ws.cell(row=row, column=15, value=ann_ret / 100)
        cell15.font          = _font(9, bold=True, color=ann_txt)
        cell15.fill          = _fill(ann_bg)
        cell15.alignment     = _align("center")
        cell15.border        = _border()
        cell15.number_format = "0.00%"
        ws.row_dimensions[row].height = 18

    # Totals row
    totals_row = header_row + 1 + len(all_years)
    ws.cell(row=totals_row, column=1, value="TOTAL").font      = _font(10, bold=True, color=C["col_text"])
    ws.cell(row=totals_row, column=1).fill      = _fill(C["col_bg"])
    ws.cell(row=totals_row, column=1).alignment = _align("center")
    ws.cell(row=totals_row, column=1).border    = _border()

    for m_num in range(1, 13):
        col       = m_num + 1
        col_total = sum(
            float(pivot.loc[yr, m_num]) if m_num in pivot.columns else 0.0
            for yr in all_years
            if not pd.isna(pivot.loc[yr, m_num] if m_num in pivot.columns else 0.0)
        )
        bg_v, txt_v = _pnl_color(col_total)
        cell = ws.cell(row=totals_row, column=col, value=col_total)
        cell.font          = _font(9, bold=True, color=txt_v)
        cell.fill          = _fill(bg_v)
        cell.alignment     = _align("center")
        cell.border        = _border()
        cell.number_format = '#,##0.00'

    ws.row_dimensions[totals_row].height = 20

    # Column widths
    ws.column_dimensions["A"].width = 10
    for c in range(2, 16):
        ws.column_dimensions[get_column_letter(c)].width = 14


def _build_trade_log(wb: Workbook, trades_df: pd.DataFrame):
    """
    Sheet 5 — Full Trade Log
    Every trade with complete entry/exit detail, partial closes, and reason.
    Rows alternate light/dark. Wins = green tint, Losses = red tint.
    """
    ws = wb.create_sheet("5. Trade Log")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"   # Freeze header rows

    _merge_write(ws, "A1:W1",
                 "FULL TRADE LOG",
                 font=_font(13, bold=True, color=C["header_text"]),
                 fill=_fill(C["header_bg"]),
                 alignment=_align("center"))
    ws.row_dimensions[1].height = 28

    columns = [
        ("#",               8),
        ("Instrument",      16),
        ("Direction",       11),
        ("Lots",            8),
        ("Entry Price",     14),
        ("Exit Price",      14),
        ("Open Time",       20),
        ("Close Time",      20),
        ("SL Price",        14),
        ("TP1 Price",       14),
        ("TP2 Price",       14),
        ("TP3 Price",       14),
        ("SL Pips",         10),
        ("Pips P&L",        10),
        ("Gross P&L ($)",   15),
        ("Commission ($)",  15),
        ("Net P&L ($)",     14),
        ("Exit Reason",     16),
        ("Score",           8),
        ("Regime",          11),
        ("TP1 Hit",         9),
        ("TP2 Hit",         9),
        ("Partials Detail", 40),
    ]

    col_keys = [
        "trade_id", "instrument", "direction", "lots",
        "entry_price", "exit_price", "open_time", "close_time",
        "sl_price", "tp1_price", "tp2_price", "tp3_price",
        "sl_pips", "pips_pnl", "gross_pnl", "commission",
        "net_pnl", "exit_reason", "score", "regime",
        "tp1_hit", "tp2_hit", "partials_detail",
    ]

    # Column headers
    for col_idx, (col_name, col_w) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font      = _font(9, bold=True, color=C["col_text"])
        cell.fill      = _fill(C["col_bg"])
        cell.alignment = _align("center")
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w
    ws.row_dimensions[2].height = 20

    # Data rows
    df = trades_df.copy()
    df["open_time"]  = pd.to_datetime(df["open_time"]).dt.strftime("%Y-%m-%d %H:%M")
    df["close_time"] = pd.to_datetime(df["close_time"]).dt.strftime("%Y-%m-%d %H:%M") \
                       if "close_time" in df.columns else "—"

    for row_idx, row_data in enumerate(df.itertuples(index=False), 3):
        net_pnl = getattr(row_data, "net_pnl", 0) or 0
        bg_row  = C["pos_bg"] if net_pnl > 0 else C["neg_bg"] if net_pnl < 0 else \
                  (C["row_light"] if row_idx % 2 == 0 else C["row_dark"])

        for col_idx, key in enumerate(col_keys, 1):
            val = getattr(row_data, key, "—") if hasattr(row_data, key) else "—"
            if val is None or (isinstance(val, float) and np.isnan(val)):
                val = "—"

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = _fill(bg_row)
            cell.border    = _border()
            cell.alignment = _align("center")
            cell.font      = _font(8.5)

            # Number formatting
            if key in ("entry_price", "exit_price", "sl_price",
                       "tp1_price", "tp2_price", "tp3_price"):
                if isinstance(val, (int, float)):
                    cell.number_format = "#,##0.00"
            elif key in ("gross_pnl", "commission", "net_pnl"):
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
                    txt = C["pos_text"] if val > 0 else C["neg_text"] if val < 0 else "000000"
                    cell.font = _font(8.5, bold=(key == "net_pnl"), color=txt)
            elif key == "score":
                if isinstance(val, (int, float)):
                    cell.number_format = "0.0"
            elif key in ("tp1_hit", "tp2_hit"):
                cell.value = "✓" if val else "—"
                if val: cell.font = _font(8.5, color=C["pos_text"], bold=True)

        ws.row_dimensions[row_idx].height = 15


def _build_equity_data(wb: Workbook, trades_df: pd.DataFrame):
    """
    Sheet 6 — Equity Curve Data
    Columns: Date, XAU_USD, SPX500_USD, DE30_EUR, US30_USD, NAS100_USD, Portfolio
    """
    ws = wb.create_sheet("6. Equity Data")
    ws.sheet_view.showGridLines = False

    _merge_write(ws, "A1:H1",
                 "EQUITY CURVE TIME-SERIES DATA",
                 font=_font(13, bold=True, color=C["header_text"]),
                 fill=_fill(C["header_bg"]),
                 alignment=_align("center"))
    ws.row_dimensions[1].height = 28

    instruments = list(INSTRUMENT_CONFIG.keys())
    headers = ["Date"] + instruments + ["Portfolio"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = _font(10, bold=True, color=C["col_text"])
        cell.fill      = _fill(C["col_bg"])
        cell.alignment = _align("center")
        cell.border    = _border()
    ws.row_dimensions[2].height = 18

    df = trades_df.copy()
    df["open_time"] = pd.to_datetime(df["open_time"])
    df["date"]      = df["open_time"].dt.date

    # Daily equity per instrument
    all_dates = sorted(df["date"].unique())
    cum_equity = {sym: INITIAL_CAPITAL for sym in instruments}
    cum_portfolio = INITIAL_CAPITAL

    for row_idx, dt in enumerate(all_dates, 3):
        day_df = df[df["date"] == dt]
        row_vals = [dt.strftime("%Y-%m-%d")]

        for sym in instruments:
            sym_day = day_df[day_df["instrument"] == sym]
            daily_pnl = sym_day["net_pnl"].sum() if not sym_day.empty else 0.0
            cum_equity[sym] = round(cum_equity[sym] + daily_pnl, 2)
            row_vals.append(cum_equity[sym])

        port_pnl = day_df["net_pnl"].sum() if not day_df.empty else 0.0
        cum_portfolio = round(cum_portfolio + port_pnl, 2)
        row_vals.append(cum_portfolio)

        bg = C["row_light"] if row_idx % 2 == 0 else C["row_dark"]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = _fill(bg)
            cell.alignment = _align("center")
            cell.border    = _border()
            cell.font      = _font(8.5)
            if col_idx > 1:
                cell.number_format = "#,##0.00"

        ws.row_dimensions[row_idx].height = 14

    ws.column_dimensions["A"].width = 16
    for i in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18


# ═══════════════════════════════════════════════════════════════════════════
# MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════

def generate_report():
    os.makedirs(REPORTS_DIR, exist_ok=True)

    pkl_path = os.path.join(RESULTS_DIR, "backtest_results.pkl")
    if not os.path.exists(pkl_path):
        logger.error(
            f"Results file not found: {pkl_path}\n"
            f"  → Run Step 2 first:  python -m scripts.02_run_backtest"
        )
        return

    with open(pkl_path, "rb") as f:
        results = pickle.load(f)

    trades_df  = pd.DataFrame(results.get("trades", []))
    metrics    = results.get("metrics", {})
    start_date = results.get("backtest_start", "—")
    end_date   = results.get("backtest_end", "—")

    if trades_df.empty:
        logger.error("No trades in results file. Cannot generate report.")
        return

    logger.info("=" * 60)
    logger.info("  APEX SCALPER — STEP 4: EXCEL REPORT")
    logger.info("=" * 60)

    wb = Workbook()

    logger.info("  [1/6] Sheet 1 — Summary Dashboard ...")
    _build_summary(wb, metrics, trades_df, start_date, end_date)

    logger.info("  [2/6] Sheet 2 — Instrument Performance ...")
    _build_instrument_performance(wb, metrics)

    logger.info("  [3/6] Sheet 3 — Monthly Returns (Portfolio) ...")
    _build_monthly_heatmap(wb, trades_df, metrics,
                           sheet_num=3,
                           sheet_title="3. Monthly Returns (Portfolio)",
                           symbol=None)

    logger.info("  [4/6] Sheet 4 — Monthly Returns (Per Asset) ...")
    ws_assets = wb.create_sheet("4. Monthly Returns (Assets)")
    ws_assets.sheet_view.showGridLines = False
    # Create one heatmap per instrument on the same sheet, stacked vertically
    # We delegate to a separate workbook temporarily and copy, or just run per-instrument
    # For cleanliness: create separate sheet for each instrument using a combined layout
    start_row = 1
    for symbol in INSTRUMENT_CONFIG.keys():
        inst_trades = trades_df[trades_df["instrument"] == symbol]
        if inst_trades.empty:
            continue
        # Build heatmap block in a temp wb, then we'll note it
        _build_monthly_heatmap(wb, inst_trades, metrics,
                               sheet_num=4,
                               sheet_title=f"4. Returns {INSTRUMENT_CONFIG[symbol]['display_name'][:8]}",
                               symbol=symbol)

    logger.info("  [5/6] Sheet 5 — Trade Log ...")
    _build_trade_log(wb, trades_df)

    logger.info("  [6/6] Sheet 6 — Equity Data ...")
    _build_equity_data(wb, trades_df)

    # Remove the placeholder sheet 4 created automatically
    if "4. Monthly Returns (Assets)" in wb.sheetnames:
        del wb["4. Monthly Returns (Assets)"]

    # Save
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"APEX_Scalper_Report_{start_date}_{end_date}_{ts}.xlsx"
    out_path = os.path.join(REPORTS_DIR, filename)
    wb.save(out_path)

    logger.info("=" * 60)
    logger.info("  REPORT COMPLETE")
    logger.info(f"  Saved: {out_path}")
    logger.info("=" * 60)


if __name__ == "__main__":
    generate_report()
